"""
üìä Export Service - Xu·∫•t danh s√°ch h√≥a ƒë∆°n
==========================================

H·ªó tr·ª£ export:
‚úÖ Excel (.xlsx)
‚úÖ PDF (.pdf)
‚úÖ CSV (.csv)
‚úÖ JSON (.json)

C√≥ th·ªÉ filter theo ng√†y, th√°ng, nƒÉm
"""

import logging
from typing import List, Dict, Any, Optional
from datetime import datetime, timedelta
import json
import csv
import io
from pathlib import Path

logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("‚ö†Ô∏è openpyxl not available - Excel export disabled")

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    logger.warning("‚ö†Ô∏è reportlab not available - PDF export disabled")


class ExportService:
    """Service xu·∫•t danh s√°ch h√≥a ƒë∆°n"""
    
    def __init__(self, db_tools=None):
        self.db_tools = db_tools
    
    # ===================== FILTER FUNCTIONS =====================
    
    @staticmethod
    def filter_by_date(invoices: List[Dict], specific_date: str) -> List[Dict]:
        """Filter h√≥a ƒë∆°n theo m·ªôt ng√†y c·ª• th·ªÉ"""
        target_date = datetime.strptime(specific_date, "%Y-%m-%d").date()
        
        filtered = []
        for inv in invoices:
            try:
                created_str = str(inv.get('created_at', ''))
                inv_date = datetime.fromisoformat(created_str.replace('Z', '+00:00')).date()
                if inv_date == target_date:
                    filtered.append(inv)
            except:
                pass
        
        logger.info(f"üìÖ Filtered {len(filtered)} invoices for date: {specific_date}")
        return filtered
    
    @staticmethod
    def filter_by_month(invoices: List[Dict], year: int, month: int) -> List[Dict]:
        """Filter h√≥a ƒë∆°n theo th√°ng"""
        filtered = []
        for inv in invoices:
            try:
                created_str = str(inv.get('created_at', ''))
                inv_datetime = datetime.fromisoformat(created_str.replace('Z', '+00:00'))
                if inv_datetime.year == year and inv_datetime.month == month:
                    filtered.append(inv)
            except:
                pass
        
        logger.info(f"üìÖ Filtered {len(filtered)} invoices for {year}-{month:02d}")
        return filtered
    
    @staticmethod
    def filter_by_date_range(invoices: List[Dict], start_date: str, end_date: str) -> List[Dict]:
        """Filter h√≥a ƒë∆°n trong kho·∫£ng th·ªùi gian"""
        start = datetime.strptime(start_date, "%Y-%m-%d").date()
        end = datetime.strptime(end_date, "%Y-%m-%d").date()
        
        filtered = []
        for inv in invoices:
            try:
                created_str = str(inv.get('created_at', ''))
                inv_date = datetime.fromisoformat(created_str.replace('Z', '+00:00')).date()
                if start <= inv_date <= end:
                    filtered.append(inv)
            except:
                pass
        
        logger.info(f"üìÖ Filtered {len(filtered)} invoices from {start_date} to {end_date}")
        return filtered
    
    # ===================== EXPORT TO CSV =====================
    
    @staticmethod
    def export_to_csv(invoices: List[Dict]) -> str:
        """Xu·∫•t h√≥a ƒë∆°n ra CSV"""
        try:
            output = io.StringIO()
            
            if not invoices:
                logger.warning("‚ö†Ô∏è No invoices to export")
                return ""
            
            # Get field names t·ª´ first invoice
            fieldnames = list(invoices[0].keys())
            
            writer = csv.DictWriter(output, fieldnames=fieldnames)
            writer.writeheader()
            
            for inv in invoices:
                writer.writerow(inv)
            
            csv_content = output.getvalue()
            logger.info(f"‚úÖ Exported {len(invoices)} invoices to CSV ({len(csv_content)} bytes)")
            
            return csv_content
        
        except Exception as e:
            logger.error(f"‚ùå CSV export error: {e}")
            return ""
    
    # ===================== EXPORT TO JSON =====================
    
    @staticmethod
    def export_to_json(invoices: List[Dict], pretty: bool = True) -> str:
        """Xu·∫•t h√≥a ƒë∆°n ra JSON"""
        try:
            if pretty:
                json_content = json.dumps(invoices, indent=2, ensure_ascii=False)
            else:
                json_content = json.dumps(invoices, ensure_ascii=False)
            
            logger.info(f"‚úÖ Exported {len(invoices)} invoices to JSON ({len(json_content)} bytes)")
            return json_content
        
        except Exception as e:
            logger.error(f"‚ùå JSON export error: {e}")
            return ""
    
    # ===================== EXPORT TO EXCEL =====================
    
    @staticmethod
    def export_to_excel(invoices: List[Dict]) -> bytes:
        """Xu·∫•t h√≥a ƒë∆°n ra Excel"""
        if not OPENPYXL_AVAILABLE:
            logger.error("‚ùå openpyxl not installed")
            return b""
        
        try:
            from openpyxl import Workbook
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Invoices"
            
            # ƒê·ªãnh d·∫°ng header
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            if not invoices:
                logger.warning("‚ö†Ô∏è No invoices to export")
                return b""
            
            # Headers
            headers = list(invoices[0].keys())
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
            
            # Data
            for row_num, invoice in enumerate(invoices, 2):
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = invoice.get(header, "")
                    cell.border = border
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # Auto-adjust column width
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            excel_bytes = output.getvalue()
            
            logger.info(f"‚úÖ Exported {len(invoices)} invoices to Excel ({len(excel_bytes)} bytes)")
            return excel_bytes
        
        except Exception as e:
            logger.error(f"‚ùå Excel export error: {e}")
            return b""
    
    # ===================== EXPORT TO PDF =====================
    
    @staticmethod
    def export_to_pdf(invoices: List[Dict]) -> bytes:
        """Xu·∫•t h√≥a ƒë∆°n ra PDF"""
        if not REPORTLAB_AVAILABLE:
            logger.error("‚ùå reportlab not installed")
            return b""
        
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            
            output = io.BytesIO()
            doc = SimpleDocTemplate(output, pagesize=letter)
            story = []
            
            # Title
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                textColor=colors.HexColor('#4472C4'),
                spaceAfter=12,
                alignment=1
            )
            
            title = Paragraph("üìã Danh s√°ch h√≥a ƒë∆°n", title_style)
            story.append(title)
            story.append(Spacer(1, 0.2 * inch))
            
            # Date
            date_style = ParagraphStyle(
                'DateStyle',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.grey,
                spaceAfter=12
            )
            date_text = Paragraph(f"Ng√†y xu·∫•t: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", date_style)
            story.append(date_text)
            story.append(Spacer(1, 0.2 * inch))
            
            # Table
            if invoices:
                headers = list(invoices[0].keys())
                
                # Gi·ªõi h·∫°n columns ƒë·ªÉ fit v√†o PDF
                data = [[str(h) for h in headers[:8]]]  # Max 8 columns
                for inv in invoices:
                    row = [str(inv.get(h, ""))[:50] for h in headers[:8]]  # Max 50 chars per cell
                    data.append(row)
                
                table = Table(data, repeatRows=1)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F0F0')]),
                ]))
                
                story.append(table)
            
            # Summary
            story.append(Spacer(1, 0.2 * inch))
            summary_style = ParagraphStyle(
                'SummaryStyle',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.grey
            )
            summary = Paragraph(f"<b>T·ªïng c·ªông:</b> {len(invoices)} h√≥a ƒë∆°n", summary_style)
            story.append(summary)
            
            # Build PDF
            doc.build(story)
            pdf_bytes = output.getvalue()
            
            logger.info(f"‚úÖ Exported {len(invoices)} invoices to PDF ({len(pdf_bytes)} bytes)")
            return pdf_bytes
        
        except Exception as e:
            logger.error(f"‚ùå PDF export error: {e}")
            return b""
    
    # ===================== SUMMARY & STATISTICS =====================
    
    @staticmethod
    def calculate_statistics(invoices: List[Dict]) -> Dict[str, Any]:
        """T√≠nh to√°n th·ªëng k√™"""
        if not invoices:
            return {}
        
        total_amount = 0
        total_confidence = 0
        invoice_types = {}
        
        for inv in invoices:
            # Total amount
            try:
                amount_str = str(inv.get('total_amount', '0')).replace(',', '').replace(' VND', '')
                total_amount += float(amount_str)
            except:
                pass
            
            # Confidence
            try:
                total_confidence += float(inv.get('confidence_score', 0))
            except:
                pass
            
            # Invoice types
            inv_type = inv.get('invoice_type', 'Unknown')
            invoice_types[inv_type] = invoice_types.get(inv_type, 0) + 1
        
        avg_confidence = total_confidence / len(invoices) if invoices else 0
        
        return {
            'total_invoices': len(invoices),
            'total_amount': total_amount,
            'average_confidence': round(avg_confidence, 2),
            'invoice_types': invoice_types
        }


# Global instance
_export_service = None

def get_export_service(db_tools=None):
    """Get ExportService singleton"""
    global _export_service
    if _export_service is None:
        _export_service = ExportService(db_tools)
    return _export_service
